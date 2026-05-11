"""
J.A.R.V.I.S. — Just A Rather Very Intelligent System
Backend principal — Flask + SocketIO + Gemini + Groq
VERSÃO 2.3 — Consciência Avançada

Novidades v2.3:
  - Wake Word offline (Whisper tiny)
  - Confirmação vocal para ações críticas
  - Briefing diário automático
  - Orb reativo à emoção (emotion_engine)
  - Modo Pomodoro por voz
  - Humor contextual (sextas, segundas, feriados)
  - Self-reflection loop semanal
  - System prompt centralizado em neural_core
"""

import os
import sys
import json
import base64
import asyncio
import datetime
import subprocess
import webbrowser
import platform
import time
import tempfile
from pathlib import Path
from threading import Thread

# Garante UTF-8 no console Windows (evita UnicodeEncodeError no banner)
if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_socketio import SocketIO, emit
import requests
import psutil

from jarvis_config import load_environment, load_settings

# ─── CARREGA .env ─────────────────────────────────────────────────────────────
loaded_env = load_environment()
if loaded_env:
    print(f'[JARVIS] Carregado: {loaded_env}')

# ─── IMPORTA MOTOR DE MEMÓRIA ─────────────────────────────────────────────────
from scheduler import (
    init_scheduler,
    add_task,
    cancel_task,
    get_all_tasks,
    parse_schedule_command,
)

from personalities import (
    get_personality,
    get_voice,
    get_all_personalities,
    set_personality,
    get_current_name,
    detect_personality_change,
)

from memory import (
    build_system_prompt,
    log_message,
    add_memory,
    get_memory_summary,
    extract_facts_from_message,
    clear_all_memories,
    is_profile_complete,
    set_profile,
    get_profile_field,
)

# ─── IMPORTA TEMPORAL LOBE ────────────────────────────────────────────────────
from nexus import (
    complete_setup_step,
    export_profile,
    finalize_setup,
    get_dashboard_status,
    get_plugins_summary,
    get_setup_state,
    import_profile,
    init_nexus,
    run_health_check,
    toggle_plugin,
)

try:
    import temporal_lobe as _tl
    _tl.init_temporal_tables()   # garante que episodes/reminders existam no DB
    TEMPORAL_AVAILABLE = True
    print('[JARVIS] Temporal Lobe: ATIVO')
except Exception as _e:
    TEMPORAL_AVAILABLE = False
    _tl = None
    print(f'[JARVIS] Temporal Lobe: INATIVO ({_e})')

# ─── MÓDULOS v2.3 — Consciência Avançada ─────────────────────────────────────

# Emotion Engine — detecta sentimento e reage no Orb
try:
    from emotion_engine import emit_emotion_event, get_humor_injection
    EMOTION_AVAILABLE = True
    print('[JARVIS] Emotion Engine: ATIVO')
except Exception as _ee:
    EMOTION_AVAILABLE = False
    print(f'[JARVIS] Emotion Engine: INATIVO ({_ee})')
    def emit_emotion_event(text, socketio_instance, sid): return {}
    def get_humor_injection(emotion): return ''

# Briefing diário
try:
    from briefing import generate_briefing, is_briefing_request
    BRIEFING_AVAILABLE = True
    print('[JARVIS] Briefing Engine: ATIVO')
except Exception as _be:
    BRIEFING_AVAILABLE = False
    print(f'[JARVIS] Briefing Engine: INATIVO ({_be})')
    def generate_briefing(deps): return 'Briefing indisponível no momento, Senhor.'
    def is_briefing_request(text): return False

# Confirmação vocal
try:
    from confirmation_system import (
        needs_confirmation,
        request_vocal_confirmation,
        resolve_confirmation,
        _parse_confirmation as _parse_vocal,
    )
    CONFIRM_V23_AVAILABLE = True
    print('[JARVIS] Confirmation System v2.3: ATIVO')
except Exception as _ce:
    CONFIRM_V23_AVAILABLE = False
    print(f'[JARVIS] Confirmation System v2.3: INATIVO ({_ce})')

# Pomodoro
try:
    from pomodoro import (
        init_pomodoro,
        is_pomodoro_command,
        handle_pomodoro_command,
        get_pomodoro_manager,
    )
    POMODORO_AVAILABLE = True
    print('[JARVIS] Pomodoro Engine: ATIVO')
except Exception as _pe:
    POMODORO_AVAILABLE = False
    print(f'[JARVIS] Pomodoro Engine: INATIVO ({_pe})')
    def is_pomodoro_command(text): return False
    def handle_pomodoro_command(text, manager, sid=''): return ''

# Humor Contextual
try:
    from humor_contextual import get_humor_context
    HUMOR_AVAILABLE = True
    print('[JARVIS] Humor Contextual: ATIVO')
except Exception as _he:
    HUMOR_AVAILABLE = False
    print(f'[JARVIS] Humor Contextual: INATIVO ({_he})')
    def get_humor_context(): return ''

# Wake Word
try:
    from wake_word import init_wake_word, get_wake_word_status
    WAKE_WORD_AVAILABLE = True
    print('[JARVIS] Wake Word Engine: CARREGADO (inicia após socketio)')
except Exception as _wwe:
    WAKE_WORD_AVAILABLE = False
    print(f'[JARVIS] Wake Word Engine: INATIVO ({_wwe})')
    def init_wake_word(*a, **kw): return None
    def get_wake_word_status(): return {'available': False, 'running': False}

# Self Reflection
try:
    from self_reflection import init_self_reflection, get_reflector
    REFLECTION_AVAILABLE = True
    print('[JARVIS] Self Reflection: CARREGADO (inicia após AI)')
except Exception as _re:
    REFLECTION_AVAILABLE = False
    print(f'[JARVIS] Self Reflection: INATIVO ({_re})')
    def init_self_reflection(*a, **kw): return None
    def get_reflector(): return None

# ─── PC AGENT ─────────────────────────────────────────────────────────────────
try:
    from pc_agent import (
        run_pc_agent,
        quick_screen_analysis,
        capture_screen_b64,
        get_agent_status,
    )
    PC_AGENT_AVAILABLE = True
    print('[JARVIS] PC Agent: ATIVO')
except ImportError as _pca_e:
    PC_AGENT_AVAILABLE = False
    print(f'[JARVIS] PC Agent: INATIVO ({_pca_e})')
    def run_pc_agent(*a, **kw): return {'success': False, 'error': 'PC Agent indisponível'}
    def quick_screen_analysis(*a, **kw): return 'PC Agent indisponível'
    def capture_screen_b64(*a, **kw): raise RuntimeError('PC Agent indisponível')
    def get_agent_status(): return {'available': False, 'pyautogui': False, 'pillow': False}

settings = load_settings()

app = Flask(__name__)
app.config['SECRET_KEY'] = settings.secret_key
socketio = SocketIO(app, async_mode='threading', cors_allowed_origins='*')

# ─── PERSONALIDADE DO JARVIS v2.3 ────────────────────────────────────────────
# O prompt central agora vive em neural_core.JARVIS_SYSTEM_PROMPT_V23.
# Use get_system_prompt() para montar o prompt completo com contextos injetados.
from neural_core import JARVIS_SYSTEM_PROMPT_V23, get_system_prompt as _build_system_prompt

# Alias de compatibilidade — mantém código legado funcionando sem alterações
JARVIS_SYSTEM_PROMPT = JARVIS_SYSTEM_PROMPT_V23

# ─── CONFIGURAÇÃO DE APIS ─────────────────────────────────────────────────────
GEMINI_API_KEY     = settings.gemini_api_key
GROQ_API_KEY       = settings.groq_api_key
TELEGRAM_BOT_TOKEN = settings.telegram_bot_token
TELEGRAM_CHAT_ID   = settings.telegram_chat_id

# ── Trello ──
TRELLO_API_KEY   = settings.trello_api_key
TRELLO_TOKEN     = settings.trello_token
TRELLO_BOARD_ID  = settings.trello_board_id

# ── Asana ──
ASANA_TOKEN      = settings.asana_token
ASANA_PROJECT_ID = settings.asana_project_id

# ── NewsAPI ──
NEWS_API_KEY  = settings.news_api_key
NEWS_TOPICS   = settings.news_topics

# ── Motor de Voz + ElevenLabs ──
MOTOR_VOZ           = settings.motor_voz  # opcoes: elevenlabs | xtts | edge
ELEVENLABS_API_KEY  = settings.elevenlabs_api_key
ELEVENLABS_VOICE_ID = settings.elevenlabs_voice_id  # voz padrão (JARVIS)
ELEVENLABS_MODEL    = settings.elevenlabs_model

# Ajuste fino de voz — edite conforme seu gosto (0.0 a 1.0)
ELEVENLABS_STABILITY        = settings.elevenlabs_stability
ELEVENLABS_SIMILARITY_BOOST = settings.elevenlabs_similarity_boost
ELEVENLABS_STYLE            = settings.elevenlabs_style
ELEVENLABS_SPEED            = settings.elevenlabs_speed

# Mapa de personalidade → voice_id
# Deixe vazio para usar a voz padrão nessa personalidade
ELEVENLABS_VOICE_MAP = settings.elevenlabs_voice_map

ELEVENLABS_AVAILABLE = bool(ELEVENLABS_API_KEY and ELEVENLABS_VOICE_ID)
print(f'[JARVIS] ElevenLabs: {"ATIVO" if ELEVENLABS_AVAILABLE else "INATIVO"}')
if ELEVENLABS_AVAILABLE:
    mapped = [k for k, v in ELEVENLABS_VOICE_MAP.items() if v]
    print(f'[JARVIS] Vozes mapeadas: {mapped if mapped else ["padrão para todas"]}')

# ── LiveKit ──
LIVEKIT_URL        = settings.livekit_url
LIVEKIT_API_KEY    = settings.livekit_api_key
LIVEKIT_API_SECRET = settings.livekit_api_secret
LIVEKIT_AVAILABLE  = bool(LIVEKIT_URL and LIVEKIT_API_KEY and LIVEKIT_API_SECRET)
print(f'[JARVIS] LiveKit: {"ATIVO" if LIVEKIT_AVAILABLE else "INATIVO"}')

# ── Mem0 ──
MEM0_API_KEY   = settings.mem0_api_key
MEM0_USER_ID   = settings.mem0_user_id
MEM0_AVAILABLE = False
mem0_client    = None

if MEM0_API_KEY:
    try:
        from mem0 import MemoryClient
        mem0_client    = MemoryClient(api_key=MEM0_API_KEY)
        MEM0_AVAILABLE = True
        print('[JARVIS] Mem0: ATIVO')
    except Exception as e:
        print(f'[JARVIS] Mem0 não disponível: {e}')
else:
    print('[JARVIS] Mem0: INATIVO (MEM0_API_KEY não configurada)')

print(f'[JARVIS] GEMINI_API_KEY presente: {bool(GEMINI_API_KEY)}')
print(f'[JARVIS] GROQ_API_KEY presente:   {bool(GROQ_API_KEY)}')

GEMINI_MODELS = [
    'gemini-2.0-flash',
    'gemini-2.0-flash-lite',
]

# ── Groq — PRIMÁRIO v2.3 (máxima prioridade para chat e intent) ──
try:
    from groq import Groq
    groq_client    = Groq(api_key=GROQ_API_KEY) if GROQ_API_KEY else None
    GROQ_AVAILABLE = bool(GROQ_API_KEY and groq_client)
except Exception as e:
    print(f'[JARVIS] Groq não disponível: {e}')
    groq_client    = None
    GROQ_AVAILABLE = False

# ── Gemini — APENAS VISÃO (screenshots, imagens, webcam) ──
# Nunca usado para texto puro — reservado exclusivamente para multimodal
try:
    from google import genai as google_genai
    gemini_client    = google_genai.Client(api_key=GEMINI_API_KEY) if GEMINI_API_KEY else None
    GEMINI_AVAILABLE = bool(GEMINI_API_KEY and gemini_client)
except Exception as e:
    print(f'[JARVIS] Gemini não disponível: {e}')
    gemini_client    = None
    GEMINI_AVAILABLE = False

# ── Ollama — FALLBACK OFFLINE (apenas se Groq falhar completamente) ──
OLLAMA_BASE_URL     = settings.ollama_base_url
OLLAMA_MODEL        = settings.ollama_model
OLLAMA_VISION_MODEL = settings.ollama_vision_model
try:
    _ollama_test     = requests.get(f'{OLLAMA_BASE_URL}/api/tags', timeout=3)
    OLLAMA_AVAILABLE = _ollama_test.status_code == 200
except Exception:
    OLLAMA_AVAILABLE = False

# mantidos para compatibilidade com módulos de skill
cerebras_client    = None
CEREBRAS_AVAILABLE = False

print(f'[JARVIS] v2.3 LLM Stack: Groq {"ON" if GROQ_AVAILABLE else "OFF"} (primário) | '
      f'Gemini {"ON" if GEMINI_AVAILABLE else "OFF"} (visão) | '
      f'Ollama {"ON" if OLLAMA_AVAILABLE else "OFF"} (offline fallback)')

# ── LLM Router v2.3 — inicializado após todos os clientes ──────────────────
from llm_router import init_router as _init_router, get_router
_llm_router = _init_router(
    groq_client          = groq_client,
    gemini_client        = gemini_client,
    ollama_url           = OLLAMA_BASE_URL,
    ollama_model         = OLLAMA_MODEL,
    ollama_vision_model  = OLLAMA_VISION_MODEL,
    ollama_available     = OLLAMA_AVAILABLE,
    gemini_available     = GEMINI_AVAILABLE,
)
print('[JARVIS] LLM Router v2.3: ATIVO')

# Sessões de chat em memória (contexto da sessão atual)
chat_sessions = {}

# ─── APLICATIVOS ──────────────────────────────────────────────────────────────
APPS_WINDOWS = {
    "calculadora": "calc.exe", "calculator": "calc.exe",
    "notepad": "notepad.exe", "bloco de notas": "notepad.exe",
    "chrome": "chrome.exe", "google chrome": "chrome.exe",
    "firefox": "firefox.exe", "explorador": "explorer.exe",
    "explorer": "explorer.exe", "paint": "mspaint.exe",
    "cmd": "cmd.exe", "terminal": "cmd.exe",
    "prompt de comando": "cmd.exe", "spotify": "spotify.exe",
    "discord": "discord.exe", "vscode": "code.exe",
    "visual studio code": "code.exe", "word": "WINWORD.EXE",
    "excel": "EXCEL.EXE", "powerpoint": "POWERPNT.EXE",
    "steam": "steam.exe", "obs": "obs64.exe",
    "vlc": "vlc.exe", "telegram": "telegram.exe",
    "whatsapp": "WhatsApp.exe",
    "spotify desktop": "start spotify",
    "spotify web": "start spotify",
    "spotify app": "start spotify",
}

APPS_LINUX = {
    "calculadora": "gnome-calculator", "calculator": "gnome-calculator",
    "notepad": "gedit", "bloco de notas": "gedit",
    "chrome": "google-chrome", "google chrome": "google-chrome",
    "firefox": "firefox", "explorador": "nautilus",
    "explorer": "nautilus", "terminal": "gnome-terminal",
    "spotify": "spotify", "discord": "discord",
    "vscode": "code", "visual studio code": "code",
    "vlc": "vlc", "telegram": "telegram-desktop",
}

def get_app_command(app_name: str):
    name_lower = app_name.lower().strip()
    if platform.system() == 'Windows':
        return APPS_WINDOWS.get(name_lower)
    return APPS_LINUX.get(name_lower)

from skills.intent_engine import detect_intent
from skills.system import execute_open_app, execute_search_web, execute_open_youtube, execute_system_info, execute_manage_files, execute_type_text, execute_open_url, execute_get_clipboard
from skills.media import execute_spotify_control, execute_protocol_work_time, execute_control_music
from skills.communication import execute_send_whatsapp, execute_send_telegram
from skills.productivity import execute_trello_action, execute_asana_action, execute_schedule_task
from skills.vision import execute_analyze_screen, execute_analyze_screen_quick, execute_take_screenshot, execute_analyze_media, execute_pc_agent_task
from skills.information import execute_get_weather, execute_get_news
from skills.instagram import execute_instagram_post, execute_instagram_story, execute_instagram_dm
from skills.google_workspace import execute_google_sheets, execute_google_calendar, execute_google_maps

import skills.intent_engine as _ie
import skills.system as _sys
import skills.media as _med
import skills.communication as _com
import skills.productivity as _prod
import skills.vision as _vis
import skills.information as _inf
import skills.instagram as _ig
import skills.google_workspace as _gw

def _init_skills():
    deps = {
        'emit': socketio.emit,
        'OLLAMA_AVAILABLE': OLLAMA_AVAILABLE,
        'OLLAMA_BASE_URL': OLLAMA_BASE_URL,
        'OLLAMA_MODEL': OLLAMA_MODEL,
        'OLLAMA_VISION_MODEL': OLLAMA_VISION_MODEL,
        'CEREBRAS_AVAILABLE': False,
        'cerebras_client': None,
        'GROQ_AVAILABLE': GROQ_AVAILABLE,
        'groq_client': groq_client,
        'GEMINI_AVAILABLE': GEMINI_AVAILABLE,
        'gemini_client': gemini_client,
        'GEMINI_MODELS': GEMINI_MODELS,
        'TELEGRAM_BOT_TOKEN': TELEGRAM_BOT_TOKEN,
        'TELEGRAM_CHAT_ID': TELEGRAM_CHAT_ID,
        'TRELLO_API_KEY': TRELLO_API_KEY,
        'TRELLO_TOKEN': TRELLO_TOKEN,
        'TRELLO_BOARD_ID': TRELLO_BOARD_ID,
        'ASANA_TOKEN': ASANA_TOKEN,
        'ASANA_PROJECT_ID': ASANA_PROJECT_ID,
        'NEWS_API_KEY': NEWS_API_KEY,
        'NEWS_TOPICS': NEWS_TOPICS,
        'PC_AGENT_AVAILABLE': PC_AGENT_AVAILABLE,
        'get_app_command': get_app_command,
        'get_spotify': get_spotify,
        'run_pc_agent': run_pc_agent,
        'quick_screen_analysis': quick_screen_analysis,
        'capture_screen_b64': capture_screen_b64,
        'get_all_tasks': get_all_tasks,
        'add_task': add_task,
        'PC_AGENT_SAFE_MODE': settings.pc_agent_safe_mode
    }
    _ie.setup(deps)
    _sys.setup(deps)
    _med.setup(deps)
    _com.setup(deps)
    _prod.setup(deps)
    _vis.setup(deps)
    _inf.setup(deps)
    _ig.setup(deps)
    _gw.setup(deps)



# ─── MOTOR DE INTENÇÕES ───────────────────────────────────────────────────────
INTENT_PROMPT = """Analise o texto abaixo e retorne APENAS um JSON com a intenção detectada.

Texto: "{text}"

INTENÇÕES ESPECIAIS:
- "protocol_work_time": SE o texto contiver expressões como "hora do papai trabalhar",
  "protocolo papai", "protocolo trabalho", "hora de trabalhar papai", "papai vai trabalhar".
  Protocolo especial que toca Back in Black do AC/DC no Spotify.
- "get_news": SE pedir notícias, manchetes, novidades, atualizações sobre tópicos.
  Extraia o tópico em "query" se mencionado. Exemplos: "me dê notícias de tecnologia",
  "quais as manchetes de hoje", "novidades sobre inteligência artificial".
- "trello_action": SE mencionar Trello — criar card, listar tarefas do Trello, mover card, ver board.
- "asana_action": SE mencionar Asana — criar tarefa, listar tarefas, completar tarefa no Asana.
- "pc_agent_task": SE pedir para EXECUTAR algo autonomamente no PC que requer ver a tela e agir
  (ex: "abra o notepad e escreva minha lista", "preencha esse formulário", "clique no botão de enviar
  no Chrome", "arraste esse arquivo para a pasta X"). Requer tarefa complexa com múltiplos passos.
  Extraia a tarefa completa no campo "task".
- "analyze_screen_quick": SE pedir para DESCREVER ou ANALISAR o que está na tela SEM executar ações
  (ex: "o que está na minha tela?", "qual erro apareceu?", "descreva o que está aberto",
  "leia esse texto na tela"). Extraia a pergunta em "query".

Para "trello_action" e "asana_action", extraia com precisão:
- "action": create|list|complete|move|delete
- "title": título exato da tarefa/card
- "description": descrição detalhada se houver
- "due_date": data de vencimento no formato YYYY-MM-DD se mencionada
- "priority": high|medium|low se mencionada
- "list_name": nome da lista/coluna do Trello (ex: "A fazer", "Em andamento", "Concluído")

Retorne SOMENTE este JSON, sem texto adicional, sem markdown:
{{
  "intent": "open_app|search_web|open_youtube|send_whatsapp|send_telegram|spotify_control|control_music|manage_files|system_info|analyze_screen|get_weather|open_url|type_text|take_screenshot|get_clipboard|get_news|trello_action|asana_action|protocol_work_time|pc_agent_task|analyze_screen_quick|conversation",
  "params": {{
    "app_name": "nome do app",
    "query": "texto de busca ou nome da musica/playlist/artista ou tópico de notícia ou pergunta sobre a tela",
    "task": "descrição completa da tarefa para o agente de PC executar autonomamente",
    "contact": "nome do contato",
    "phone": "+5524999979286",
    "message": "mensagem a enviar",
    "action": "play|pause|next|previous|volume_up|volume_down|mute|search|create|list|complete|move|delete",
    "volume": 50,
    "file_path": "/caminho/do/arquivo",
    "operation": "list|create|delete|open",
    "url": "https://...",
    "text": "texto a digitar",
    "city": "nome da cidade",
    "title": "título da tarefa ou card",
    "description": "descrição da tarefa",
    "due_date": "YYYY-MM-DD",
    "priority": "high|medium|low",
    "list_name": "nome da lista do Trello",
    "count": 3
  }}
}}"""


# ─── EXECUTORES DE AÇÕES ──────────────────────────────────────────────────────






# ─── SPOTIFY ──────────────────────────────────────────────────────────────────
SPOTIFY_CLIENT_ID     = os.getenv('SPOTIFY_CLIENT_ID', '')
SPOTIFY_CLIENT_SECRET = os.getenv('SPOTIFY_CLIENT_SECRET', '')
SPOTIFY_REDIRECT_URI  = os.getenv('SPOTIFY_REDIRECT_URI', 'http://127.0.0.1:8888/callback')

_spotify = None

def get_spotify():
    """Inicializa e retorna cliente Spotify autenticado (lazy init)."""
    global _spotify
    if _spotify:
        return _spotify
    if not SPOTIFY_CLIENT_ID or not SPOTIFY_CLIENT_SECRET:
        return None
    try:
        import spotipy
        from spotipy.oauth2 import SpotifyOAuth
        scope = (
            'user-read-playback-state '
            'user-modify-playback-state '
            'user-read-currently-playing '
            'streaming '
            'playlist-read-private '
            'user-library-read'
        )
        auth = SpotifyOAuth(
            client_id=SPOTIFY_CLIENT_ID,
            client_secret=SPOTIFY_CLIENT_SECRET,
            redirect_uri=SPOTIFY_REDIRECT_URI,
            scope=scope,
            open_browser=True,
            cache_path='.spotify_cache',
        )
        _spotify = spotipy.Spotify(auth_manager=auth)
        # Testa conexão
        _spotify.current_user()
        print('[JARVIS] Spotify conectado com sucesso')
        return _spotify
    except Exception as e:
        print(f'[JARVIS] Spotify erro: {e}')
        _spotify = None
        return None







# ─── VISÃO COMPUTACIONAL ──────────────────────────────────────────────────────

# ─── NOVAS FERRAMENTAS ────────────────────────────────────────────────────────






# ─── ANÁLISE DE IMAGEM/MÍDIA COM IA ──────────────────────────────────────────



# ─── NOTÍCIAS PERSONALIZADAS ──────────────────────────────────────────────────


# ─── TRELLO ───────────────────────────────────────────────────────────────────


# ─── ASANA ────────────────────────────────────────────────────────────────────


# ─── DISPATCHER ───────────────────────────────────────────────────────────────
# ─── SISTEMA DE CONFIRMAÇÃO ────────────────────────────────────────────────────
import uuid
from threading import Event

_pending_confirms = {}
_confirm_results  = {}

CONFIRM_INTENTS = {'send_telegram', 'send_whatsapp', 'manage_files'}

CONFIRM_LABELS = {
    'send_telegram': ('Enviar mensagem via Telegram', 'Uma mensagem será enviada pelo seu bot do Telegram.'),
    'send_whatsapp': ('Enviar mensagem via WhatsApp',  'Uma mensagem será enviada pelo WhatsApp Web.'),
    'manage_files':  ('Operação em arquivos',           'Uma operação de arquivo será executada no seu sistema.'),
}

def request_confirmation(intent: str, params: dict, sid: str) -> bool:
    confirm_id = str(uuid.uuid4())
    label, detail = CONFIRM_LABELS.get(intent, (intent, ''))

    if intent == 'send_telegram':
        msg = params.get('message', '')
        if msg:
            detail = f'Mensagem: "{msg[:80]}{"…" if len(msg) > 80 else ""}"'
    elif intent == 'send_whatsapp':
        phone = params.get('phone', '')
        msg   = params.get('message', '')
        detail = f'Para: {phone} | "{msg[:60]}{"…" if len(msg) > 60 else ""}"' 
    elif intent == 'manage_files':
        op   = params.get('operation', '')
        path = params.get('file_path', '')
        detail = f'Operação: {op} | Caminho: {path}'

    ev = Event()
    _pending_confirms[confirm_id] = ev
    _confirm_results[confirm_id]  = False

    socketio.emit('confirm_action', {
        'id':     confirm_id,
        'intent': intent,
        'action': label,
        'detail': detail,
    }, room=sid)

    confirmed = ev.wait(timeout=20)
    result = _confirm_results.pop(confirm_id, False)
    _pending_confirms.pop(confirm_id, None)
    return result if confirmed else False

def dispatch_intent(intent_data: dict, sid: str):
    intent = intent_data.get('intent', 'conversation')
    params = intent_data.get('params', {})
    print(f'[JARVIS] Executando: {intent} | Params: {params}')
    socketio.emit('status_update', {'step': 'executing', 'message': f'Executando: {intent}'}, room=sid)

    if intent in CONFIRM_INTENTS:
        socketio.emit('status_update', {'step': 'thinking', 'message': 'Aguardando confirmação...'}, room=sid)
        confirmed = request_confirmation(intent, params, sid)
        if not confirmed:
            return f'Ação cancelada pelo Senhor.'

    if intent == 'open_app':      return execute_open_app(params, sid)
    if intent == 'search_web':    return execute_search_web(params, sid)
    if intent == 'open_youtube':  return execute_open_youtube(params, sid)
    if intent == 'send_whatsapp': return execute_send_whatsapp(params, sid)
    if intent == 'send_telegram': return execute_send_telegram(params, sid)
    if intent == 'spotify_control': return execute_spotify_control(params, sid)
    if intent == 'control_music': return execute_control_music(params, sid)
    if intent == 'system_info':   return execute_system_info(sid)
    if intent == 'manage_files':    return execute_manage_files(params, sid)
    if intent == 'analyze_screen':  return execute_analyze_screen(params, sid)
    if intent == 'get_weather':     return execute_get_weather(params, sid)
    if intent == 'open_url':        return execute_open_url(params, sid)
    if intent == 'type_text':       return execute_type_text(params, sid)
    if intent == 'take_screenshot': return execute_take_screenshot(params, sid)
    if intent == 'get_clipboard':   return execute_get_clipboard(params, sid)
    if intent == 'analyze_media':   return execute_analyze_media(params, sid)
    if intent == 'schedule_task':   return execute_schedule_task(params, sid)
    if intent == 'get_news':        return execute_get_news(params, sid)
    if intent == 'trello_action':   return execute_trello_action(params, sid)
    if intent == 'asana_action':    return execute_asana_action(params, sid)
    if intent == 'protocol_work_time': return execute_protocol_work_time(params, sid)
    if intent == 'pc_agent_task':        return execute_pc_agent_task(params, sid)
    if intent == 'analyze_screen_quick': return execute_analyze_screen_quick(params, sid)
    if intent == 'instagram_post':    return execute_instagram_post(params, sid, image_data=_pending_files.get(sid))
    if intent == 'instagram_story':   return execute_instagram_story(params, sid, image_data=_pending_files.get(sid))
    if intent == 'instagram_dm':      return execute_instagram_dm(params, sid)
    if intent == 'google_sheets':     return execute_google_sheets(params, sid)
    if intent == 'google_calendar':   return execute_google_calendar(params, sid)
    if intent == 'google_maps':       return execute_google_maps(params, sid)
    return None

# Armazena arquivos anexados pendentes por sessão
_pending_files = {}
# Armazena contexto de documento extraído por sessão
_file_contexts = {}
# Armazena estado de pergunta pendente (fluxo conversacional)
_pending_asks = {}

# ─── EXTRAÇÃO DE FATOS VIA IA ─────────────────────────────────────────────────
FACT_PROMPT = """Analise a mensagem abaixo e extraia APENAS fatos novos e relevantes sobre o usuário.
Se não houver fatos relevantes, retorne uma lista vazia.

Mensagem: "{text}"

Retorne SOMENTE este JSON:
{{
  "facts": [
    {{"category": "pessoal|trabalho|preferencia|tecnologia|habito", "content": "fato em uma frase curta", "importance": 1}}
  ]
}}

Exemplos de fatos válidos:
- "Me chamo Pedro" → pessoal, "Usuário se chama Pedro", importance: 3
- "Trabalho como médico" → trabalho, "Trabalha como médico", importance: 2
- "Prefiro respostas curtas" → preferencia, "Prefere respostas curtas", importance: 2
- "Uso o VSCode todo dia" → tecnologia, "Usa VSCode como editor principal", importance: 1

NÃO extraia fatos sobre temas gerais, apenas sobre o USUÁRIO ESPECÍFICO."""

def extract_facts_via_ai(text: str):
    """Usa IA para extrair fatos relevantes da mensagem do usuário e salva na memória."""
    if len(text) < 10:
        return

    prompt = FACT_PROMPT.format(text=text)

    try:
        raw = None
        if GEMINI_AVAILABLE and gemini_client:
            for model in GEMINI_MODELS:
                try:
                    response = gemini_client.models.generate_content(model=model, contents=prompt)
                    raw = response.text.strip().replace('```json', '').replace('```', '').strip()
                    break
                except Exception:
                    continue

        if not raw and GROQ_AVAILABLE and groq_client:
            resp = groq_client.chat.completions.create(
                model='llama-3.3-70b-versatile',
                messages=[{'role': 'user', 'content': prompt}],
                max_tokens=200, temperature=0.1,
            )
            raw = resp.choices[0].message.content.strip().replace('```json', '').replace('```', '').strip()

        if raw:
            # Remove múltiplos JSONs ou lixo após o primeiro objeto
            raw = raw.split('\n')[0] if '\n' in raw else raw
            try:
                data = json.loads(raw)
            except json.JSONDecodeError:
                # Tenta extrair apenas o primeiro objeto JSON válido
                import re
                match = re.search(r'\{.*\}', raw, re.DOTALL)
                if match:
                    data = json.loads(match.group())
                else:
                    return
            facts = data.get('facts', [])
            for fact in facts:
                category  = fact.get('category', 'pessoal')
                content   = fact.get('content', '')
                importance = fact.get('importance', 1)
                if content and len(content) > 5:
                    add_memory(category, content, importance)
                    # Se o fato contiver nome, salva no perfil também
                    if 'se chama' in content.lower() or 'nome é' in content.lower():
                        name = content.split()[-1].strip('.,')
                        if 2 < len(name) < 30:
                            set_profile('user_name', name)
                    print(f'[JARVIS] Fato aprendido: [{category}] {content}')

    except Exception as e:
        print(f'[JARVIS] Erro ao extrair fatos: {e}')
        # Fallback para extração por regras simples
        extract_facts_from_message(text)

# ─── GERAÇÃO DE RESPOSTA IA ───────────────────────────────────────────────────
def generate_ai_response(messages: list, context: str = '', emotion: str = '') -> tuple:
    """
    Gera resposta de chat via LLM Router v2.3.
    Rota: Groq/llama-3.3-70b → Groq/qwen → Ollama → fallback.
    """
    memory_prompt = build_system_prompt()

    # Mem0
    mem0_extra = ''
    if MEM0_AVAILABLE and messages:
        ultima_msg = messages[-1].get('content', '')
        mem0_extra = mem0_search(ultima_msg, limit=5)

    # Humor contextual + emoção
    humor_ctx   = get_humor_context() if HUMOR_AVAILABLE else ''
    emotion_ctx = get_humor_injection(emotion) if (EMOTION_AVAILABLE and emotion) else ''

    # Padrões da sessão
    try:
        from neural_core import get_conditioning_context
        conditioning_ctx = get_conditioning_context()
    except Exception:
        conditioning_ctx = ''

    system = _build_system_prompt(
        memory_context=memory_prompt + mem0_extra,
        conditioning_context=conditioning_ctx,
        humor_context=humor_ctx + emotion_ctx,
        extra_context=context,
    )

    text, model_used = get_router().chat(messages, system=system)
    return text, model_used

# ─── MEM0 — MEMÓRIA DE LONGO PRAZO ───────────────────────────────────────────
def mem0_add(text: str, role: str = 'user'):
    """Salva mensagem na memória Mem0 em segundo plano."""
    if not MEM0_AVAILABLE or not mem0_client:
        return
    try:
        mem0_client.add(
            messages=[{'role': role, 'content': text}],
            user_id=MEM0_USER_ID,
        )
        print(f'[MEM0] Salvo: {text[:60]}')
    except Exception as e:
        print(f'[MEM0] Erro ao salvar: {e}')

def mem0_search(query: str, limit: int = 5) -> str:
    """Busca memórias relevantes no Mem0 para enriquecer o contexto da IA."""
    if not MEM0_AVAILABLE or not mem0_client:
        return ''
    try:
        results = mem0_client.search(query, user_id=MEM0_USER_ID, limit=limit)
        if not results:
            return ''
        memorias = [r.get('memory', '') for r in results if r.get('memory')]
        if not memorias:
            return ''
        bloco = '\n'.join(f'- {m}' for m in memorias)
        print(f'[MEM0] {len(memorias)} memórias encontradas')
        return f'\n\n## MEMÓRIAS DE LONGO PRAZO (Mem0)\n{bloco}'
    except Exception as e:
        print(f'[MEM0] Erro na busca: {e}')
        return ''

# ─── TTS ──────────────────────────────────────────────────────────────────────

# ─── GERAÇÃO DE RESPOSTA IA (STREAMING) ───────────────────────────────────────
def generate_ai_response_stream(messages: list, context: str = '', emotion: str = ''):
    """
    Streaming token-a-token via LLM Router v2.3.
    Rota: Groq/llama-3.3-70b (stream) → Ollama (stream) → fallback texto.
    """
    memory_prompt = build_system_prompt()

    mem0_extra = ''
    if MEM0_AVAILABLE and messages:
        ultima_msg = messages[-1].get('content', '')
        mem0_extra = mem0_search(ultima_msg, limit=5)

    humor_ctx   = get_humor_context() if HUMOR_AVAILABLE else ''
    emotion_ctx = get_humor_injection(emotion) if (EMOTION_AVAILABLE and emotion) else ''

    try:
        from neural_core import get_conditioning_context
        conditioning_ctx = get_conditioning_context()
    except Exception:
        conditioning_ctx = ''

    system = _build_system_prompt(
        memory_context=memory_prompt + mem0_extra,
        conditioning_context=conditioning_ctx,
        humor_context=humor_ctx + emotion_ctx,
        extra_context=context,
    )

    yield from get_router().stream(messages, system=system)

import websockets
import asyncio
import ssl as _ssl

_no_verify_ssl = _ssl.create_default_context()
_no_verify_ssl.check_hostname = False
_no_verify_ssl.verify_mode = _ssl.CERT_NONE

async def _stream_tts_elevenlabs_ws(text_generator, voice_id, emit_callback):
    uri = f"wss://api.elevenlabs.io/v1/text-to-speech/{voice_id}/stream-input?model_id=eleven_turbo_v2_5&optimize_streaming_latency=4"

    try:
        async with websockets.connect(uri, ping_interval=None, ssl=_no_verify_ssl) as websocket:
            await websocket.send(json.dumps({
                "text": " ",
                "voice_settings": {
                    "stability": ELEVENLABS_STABILITY,
                    "similarity_boost": ELEVENLABS_SIMILARITY_BOOST,
                    "style": ELEVENLABS_STYLE,
                    "speed": ELEVENLABS_SPEED,
                },
                "xi_api_key": ELEVENLABS_API_KEY
            }))

            async def listen():
                while True:
                    try:
                        message = await websocket.recv()
                        data = json.loads(message)
                        if data.get("audio"):
                            emit_callback(audio_b64=data["audio"])
                        if data.get("isFinal"):
                            break
                    except Exception as e:
                        print("[JARVIS] ElevenLabs WS Error (listen):", e)
                        break

            listen_task = asyncio.create_task(listen())

            buffer = ""
            full_text = ""
            # Pontuações que indicam fim de frase — flush nesses pontos
            # (mais responsivo que espaço, menos chamadas que palavra a palavra)
            FLUSH_CHARS = {'.', '!', '?', '\n', '—', ';'}
            FLUSH_MIN_LEN = 40  # flush só se buffer tiver conteúdo suficiente

            for token in text_generator:
                emit_callback(text_chunk=token)
                buffer += token
                full_text += token

                # Flush quando fecha uma frase ou o buffer ficou grande
                should_flush = (
                    any(c in buffer for c in FLUSH_CHARS) and len(buffer) >= FLUSH_MIN_LEN
                ) or len(buffer) >= 120

                if should_flush:
                    await websocket.send(json.dumps({"text": buffer, "try_trigger_generation": True}))
                    buffer = ""

            # Flush do restante
            if buffer:
                await websocket.send(json.dumps({"text": buffer, "try_trigger_generation": True}))

            await websocket.send(json.dumps({"text": ""}))
            await listen_task
            return full_text

    except Exception as e:
        print(f"[JARVIS] ElevenLabs WebSocket error: {e}")
        full_text = ""
        for token in text_generator:
            emit_callback(text_chunk=token)
            full_text += token
        return full_text

def run_tts_stream(text_generator, voice_id, sid):
    def emit_callback(audio_b64=None, text_chunk=None):
        if audio_b64:
            socketio.emit('audio_chunk', {'chunk_b64': audio_b64}, room=sid)
        if text_chunk:
            socketio.emit('text_chunk', {'text': text_chunk}, room=sid)

    return asyncio.run(_stream_tts_elevenlabs_ws(text_generator, voice_id, emit_callback))


@socketio.on('chat_message_stream')
def on_chat_message_stream(data):
    sid  = request.sid
    text = data.get('text', '').strip()
    if not text: return

    print(f'[JARVIS STREAM] Mensagem: {text}')

    def _process():
        try:
            socketio.emit('status_update', {'step': 'thinking', 'message': 'Analisando...'}, room=sid)

            # Prepara histórico enquanto detecta intenção em paralelo
            history = chat_sessions.get(sid, [])
            if len(history) > 12:   # reduzido de 20 → 12: menos tokens, menos latência
                history = history[-12:]
            history.append({'role': 'user', 'content': text})

            intent_data = detect_intent(text)

            action_context = None
            if intent_data.get('intent') not in ('conversation', None):
                action_context = dispatch_intent(intent_data, sid)

            # Extrai fatos em background — não bloqueia a resposta
            Thread(target=extract_facts_via_ai, args=(text,), daemon=True).start()

            socketio.emit('status_update', {'step': 'speaking', 'message': 'Respondendo...'}, room=sid)

            generator = generate_ai_response_stream(history, context=action_context or '')

            current_personality = get_current_name()
            voice_id = ELEVENLABS_VOICE_MAP.get(current_personality, ELEVENLABS_VOICE_ID)

            full_resposta = run_tts_stream(generator, voice_id, sid)

            socketio.emit('stream_end', {'intent': intent_data.get('intent', 'conversation')}, room=sid)

            history.append({'role': 'assistant', 'content': full_resposta})
            chat_sessions[sid] = history
            log_message('user', text, intent=intent_data.get('intent'))
            log_message('assistant', full_resposta)

            if MEM0_AVAILABLE:
                Thread(target=mem0_add, args=(text, 'user'), daemon=True).start()
                Thread(target=mem0_add, args=(full_resposta, 'assistant'), daemon=True).start()

        except Exception as e:
            print(f'[JARVIS] Erro no stream: {e}')
            socketio.emit('error', {'message': f'Erro no stream: {e}'}, room=sid)

    Thread(target=_process, daemon=True).start()


def generate_tts_elevenlabs(text: str):
    """
    Gera áudio TTS via ElevenLabs com a voz clonada.
    Troca automaticamente de voice_id conforme a personalidade ativa.
    Retorna base64 MP3 ou None se falhar.
    """
    if not ELEVENLABS_AVAILABLE:
        return None

    # Resolve o voice_id para a personalidade ativa
    current_personality = get_current_name()
    voice_id = ELEVENLABS_VOICE_MAP.get(current_personality, '')
    if not voice_id:
        voice_id = ELEVENLABS_VOICE_ID   # fallback para voz padrão

    try:
        import requests as req
        resp = req.post(
            f'https://api.elevenlabs.io/v1/text-to-speech/{voice_id}',
            headers={
                'xi-api-key':   ELEVENLABS_API_KEY,
                'Content-Type': 'application/json',
                'Accept':       'audio/mpeg',
            },
            json={
                'text':     text,
                'model_id': ELEVENLABS_MODEL,
                'voice_settings': {
                    'stability':         ELEVENLABS_STABILITY,
                    'similarity_boost':  ELEVENLABS_SIMILARITY_BOOST,
                    'style':             ELEVENLABS_STYLE,
                    'speed':             ELEVENLABS_SPEED,
                    'use_speaker_boost': True,
                },
            },
            timeout=12,
            verify=False,
        )

        if resp.status_code == 200:
            audio_b64 = base64.b64encode(resp.content).decode('utf-8')
            print(f'[JARVIS] TTS via ElevenLabs | personalidade: {current_personality} | voice_id: {voice_id[:8]}...')
            return audio_b64

        # 429 = quota, 401 = key inválida — loga e deixa o fallback agir
        print(f'[JARVIS] ElevenLabs {resp.status_code} — fallback ativado')
        return None

    except Exception as e:
        print(f'[JARVIS] Erro ElevenLabs: {e} — fallback ativado')
        return None


def generate_tts_edge(text: str):
    """Gera áudio TTS via edge-tts (fallback local)."""
    try:
        import edge_tts

        async def _generate():
            communicate = edge_tts.Communicate(text, voice=get_voice(), rate='-8%')
            with tempfile.NamedTemporaryFile(suffix='.mp3', delete=False) as f:
                tmp_path = f.name
            await communicate.save(tmp_path)
            with open(tmp_path, 'rb') as f:
                data = f.read()
            os.unlink(tmp_path)
            return base64.b64encode(data).decode('utf-8')

        loop = asyncio.new_event_loop()
        try:
            return loop.run_until_complete(_generate())
        finally:
            loop.close()

    except ImportError:
        print('[JARVIS] edge-tts não instalado')
        return None
    except Exception as e:
        print(f'[JARVIS] Erro edge-tts: {e}')
        return None


def generate_tts(text: str):
    """
    Gera áudio TTS com motor configurável via .env (MOTOR_VOZ):
    - elevenlabs: ElevenLabs API (PT-BR, voz clonada, troca por personalidade)
    - xtts:       XTTS local (offline, clonagem de voz)
    - edge:       edge-tts (offline, fallback garantido)
    Fallback automático para edge-tts em caso de falha.
    """
    # 1. ElevenLabs
    if MOTOR_VOZ == 'elevenlabs' and ELEVENLABS_AVAILABLE:
        result = generate_tts_elevenlabs(text)
        if result:
            return result
        print('[JARVIS] ElevenLabs falhou — próximo motor')

    # 2. XTTS local
    elif MOTOR_VOZ == 'xtts':
        try:
            from motor_voz import gerar_fala_jarvis
            result = gerar_fala_jarvis(text)
            if result:
                return result
            print('[JARVIS] XTTS falhou — próximo motor')
        except Exception as e:
            print(f'[JARVIS] Erro XTTS: {e}')

    # 3. edge-tts (sempre disponível)
    return generate_tts_edge(text)

# ─── ROTAS HTTP ───────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return send_from_directory(os.path.dirname(os.path.abspath(__file__)), 'index.html')

@app.route('/api/transcribe', methods=['POST'])
def transcribe():
    if 'audio' not in request.files:
        return jsonify({'error': 'Nenhum arquivo de áudio enviado'}), 400
    if not GROQ_AVAILABLE or not groq_client:
        return jsonify({'error': 'Groq não configurado'}), 503

    audio_file = request.files['audio']
    try:
        with tempfile.NamedTemporaryFile(suffix='.webm', delete=False) as f:
            audio_file.save(f.name)
            tmp_path = f.name

        # DIAGNÓSTICO
        size = os.path.getsize(tmp_path)
        print(f'[JARVIS] Áudio recebido: {size} bytes')

        if size < 1000:
            print(f'[JARVIS] Áudio muito pequeno — microfone não captou nada')
            os.unlink(tmp_path)
            return jsonify({'error': 'Áudio muito curto ou silencioso'}), 400

        with open(tmp_path, 'rb') as f:
            transcription = groq_client.audio.transcriptions.create(
                file=('recording.webm', f, 'audio/webm'),
                model='whisper-large-v3',
                language='pt',
            )
        os.unlink(tmp_path)
        print(f'[JARVIS] Whisper retornou: "{transcription.text}"')
        return jsonify({'text': transcription.text})
    except Exception as e:
        print(f'[JARVIS] Erro transcrição: {e}')
        return jsonify({'error': str(e)}), 500

# NOVA ROTA — retorna estado da memória para o frontend
@app.route('/api/analyze_image', methods=['POST'])
def analyze_image():
    """Recebe imagem do frontend, analisa com Gemini Vision e retorna texto + áudio."""
    if 'image' not in request.files:
        return jsonify({'error': 'Nenhuma imagem enviada'}), 400

    image_file = request.files['image']
    query = request.form.get('query', 'Descreva esta imagem em detalhes em português brasileiro.')
    sid   = request.form.get('sid', '')

    try:
        image_data = image_file.read()
        mime_type  = image_file.content_type or 'image/jpeg'

        # Armazena para uso posterior (ex: postar no Instagram)
        if sid:
            _pending_files[sid] = image_data

        params = {'query': query}
        result = execute_analyze_media(params, sid, image_data=image_data, mime_type=mime_type)

        # Gera TTS da resposta
        audio_b64 = generate_tts(result)

        return jsonify({
            'text':      result,
            'audio_b64': audio_b64,
            'success':   True,
        })
    except Exception as e:
        print(f'[JARVIS] Erro analyze_image: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/analyze_file', methods=['POST'])
def analyze_file():
    """Recebe qualquer arquivo (PDF, TXT, XLSX, imagens) e analisa."""
    if 'file' not in request.files:
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400

    uploaded = request.files['file']
    query = request.form.get('query', '')
    sid   = request.form.get('sid', '')
    filename = uploaded.filename or 'arquivo'
    ext = Path(filename).suffix.lower()

    try:
        file_data = uploaded.read()

        # Armazena para uso posterior
        if sid:
            _pending_files[sid] = file_data

        # === IMAGEM ===
        if ext in ('.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp'):
            mime_type = uploaded.content_type or 'image/jpeg'
            params = {'query': query or 'Descreva esta imagem em detalhes em português brasileiro.'}
            result = execute_analyze_media(params, sid, image_data=file_data, mime_type=mime_type)

        # === PDF ===
        elif ext == '.pdf':
            result = _extract_and_analyze_pdf(file_data, query, sid)

        # === XLSX ===
        elif ext in ('.xlsx', '.xls'):
            result = _extract_and_analyze_xlsx(file_data, query, sid)

        # === TXT / CSV ===
        elif ext in ('.txt', '.csv', '.md', '.log'):
            result = _extract_and_analyze_text(file_data, query, sid, filename)

        else:
            result = f'Formato {ext} não suportado ainda, Senhor. Formatos aceitos: PDF, TXT, XLSX, JPG/PNG.'

        # Armazena contexto para comandos futuros
        if sid and not result.startswith('Formato'):
            _file_contexts[sid] = {
                'filename': filename,
                'content_preview': result[:500],
                'ext': ext,
            }

        audio_b64 = generate_tts(result)
        return jsonify({
            'text': result,
            'audio_b64': audio_b64,
            'filename': filename,
            'success': True,
        })

    except Exception as e:
        print(f'[JARVIS] Erro analyze_file: {e}')
        return jsonify({'error': str(e)}), 500


def _extract_and_analyze_pdf(file_data: bytes, query: str, sid: str) -> str:
    """Extrai texto de PDF e analisa com IA."""
    try:
        from PyPDF2 import PdfReader
        import io
        reader = PdfReader(io.BytesIO(file_data))
        text_content = ''
        for page in reader.pages[:10]:  # Limite de 10 páginas
            text_content += page.extract_text() or ''

        if not text_content.strip():
            return 'O PDF parece estar vazio ou contém apenas imagens, Senhor.'

        num_pages = len(reader.pages)
        prompt = query or f'Resuma este documento PDF de {num_pages} páginas em português brasileiro.'
        return _analyze_text_with_ai(text_content[:3000], prompt, sid)

    except Exception as e:
        return f'Erro ao ler o PDF: {e}'


def _extract_and_analyze_xlsx(file_data: bytes, query: str, sid: str) -> str:
    """Extrai dados de XLSX e analisa com IA."""
    try:
        from openpyxl import load_workbook
        import io
        wb = load_workbook(io.BytesIO(file_data), read_only=True)
        sheets_data = []
        for sheet_name in wb.sheetnames[:3]:  # Limite de 3 abas
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(max_row=20, values_only=True):  # Limite de 20 linhas
                rows.append([str(cell) if cell is not None else '' for cell in row])
            if rows:
                sheets_data.append(f'\n=== Aba: {sheet_name} ===\n')
                for row in rows:
                    sheets_data.append(' | '.join(row))

        text_content = '\n'.join(sheets_data)
        if not text_content.strip():
            return 'A planilha parece estar vazia, Senhor.'

        prompt = query or 'Analise esta planilha e resuma seu conteúdo em português brasileiro.'
        return _analyze_text_with_ai(text_content[:3000], prompt, sid)

    except Exception as e:
        return f'Erro ao ler a planilha: {e}'


def _extract_and_analyze_text(file_data: bytes, query: str, sid: str, filename: str) -> str:
    """Analisa arquivo de texto com IA."""
    try:
        text_content = file_data.decode('utf-8', errors='ignore')
    except Exception:
        text_content = file_data.decode('latin-1', errors='ignore')

    if not text_content.strip():
        return f'O arquivo {filename} parece estar vazio, Senhor.'

    prompt = query or f'Analise o conteúdo deste arquivo ({filename}) e resuma em português brasileiro.'
    return _analyze_text_with_ai(text_content[:3000], prompt, sid)


def _analyze_text_with_ai(text: str, query: str, sid: str) -> str:
    """Envia texto extraído para IA analisar/resumir."""
    full_prompt = f"{query}\n\n--- CONTEÚDO DO DOCUMENTO ---\n{text}"

    # Ollama
    if OLLAMA_AVAILABLE:
        try:
            resp = requests.post(
                f'{OLLAMA_BASE_URL}/api/chat',
                json={
                    'model': OLLAMA_MODEL,
                    'messages': [{'role': 'user', 'content': full_prompt}],
                    'stream': False,
                    'options': {'temperature': 0.3, 'num_predict': 500},
                },
                timeout=30,
            )
            resp.raise_for_status()
            return resp.json()['message']['content']
        except Exception as e:
            print(f'[JARVIS] Ollama doc analysis falhou: {e}')

    # Gemini
    if GEMINI_AVAILABLE and gemini_client:
        for model in GEMINI_MODELS:
            try:
                response = gemini_client.models.generate_content(
                    model=model, contents=full_prompt
                )
                return response.text
            except Exception:
                continue

    # Groq
    if GROQ_AVAILABLE and groq_client:
        try:
            resp = groq_client.chat.completions.create(
                model='llama-3.3-70b-versatile',
                messages=[{'role': 'user', 'content': full_prompt}],
                max_tokens=500, temperature=0.3,
            )
            return resp.choices[0].message.content
        except Exception as e:
            print(f'[JARVIS] Groq doc analysis falhou: {e}')

    return 'Não foi possível analisar o documento no momento, Senhor.'


@app.route('/api/google/auth', methods=['GET'])
def api_google_auth():
    """Inicia fluxo OAuth do Google."""
    try:
        from skills.google_workspace import _get_google_creds, has_credentials_file
        if not has_credentials_file():
            return jsonify({
                'success': False,
                'error': 'Arquivo google_credentials.json não encontrado na pasta do Jarvis.'
            }), 400
        creds = _get_google_creds()
        if creds and creds.valid:
            return jsonify({'success': True, 'message': 'Autenticado com sucesso!'})
        return jsonify({'success': False, 'error': 'Falha na autenticação'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/google/status', methods=['GET'])
def api_google_status():
    """Verifica status da autenticação Google."""
    from skills.google_workspace import is_google_authenticated, has_credentials_file
    return jsonify({
        'authenticated': is_google_authenticated(),
        'credentials_file': has_credentials_file(),
    })

@app.route('/api/memory', methods=['GET'])
def api_memory():
    return jsonify(get_memory_summary())

# NOVA ROTA — limpa todas as memórias
@app.route('/api/memory/clear', methods=['POST'])
def api_memory_clear():
    clear_all_memories()
    return jsonify({'success': True, 'message': 'Memórias apagadas'})

# NOVA ROTA — retorna perfil do usuário
@app.route('/api/profile', methods=['GET'])
def api_profile():
    from memory import get_profile
    return jsonify(get_profile())

@app.route('/api/nexus/dashboard', methods=['GET'])
def api_nexus_dashboard():
    return jsonify(get_dashboard_status())

@app.route('/api/nexus/plugins', methods=['GET'])
def api_nexus_plugins():
    return jsonify(get_plugins_summary())

@app.route('/api/nexus/plugins/toggle', methods=['POST'])
def api_nexus_plugins_toggle():
    data = request.json or {}
    result = toggle_plugin(data.get('plugin_id', ''), bool(data.get('enabled')))
    return jsonify(result), 200 if result.get('success') else 404

@app.route('/api/nexus/health', methods=['GET'])
def api_nexus_health():
    return jsonify(run_health_check())

@app.route('/api/nexus/export', methods=['GET'])
def api_nexus_export():
    filename = f"jarvis_backup_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    output_path = os.path.join(tempfile.gettempdir(), filename)
    result = export_profile(output_path)
    if not result.get('success'):
        return jsonify(result), 500
    return send_file(output_path, as_attachment=True, download_name=filename, mimetype='application/json')

@app.route('/api/nexus/import', methods=['POST'])
def api_nexus_import():
    uploaded = request.files.get('file')
    if not uploaded:
        return jsonify({'success': False, 'error': 'Arquivo não enviado'}), 400

    temp_path = os.path.join(tempfile.gettempdir(), f"jarvis_import_{int(time.time())}.json")
    uploaded.save(temp_path)
    try:
        result = import_profile(temp_path)
    finally:
        try:
            os.unlink(temp_path)
        except OSError:
            pass
    return jsonify(result), 200 if result.get('success') else 400

@app.route('/api/nexus/setup/state', methods=['GET'])
def api_nexus_setup_state():
    return jsonify(get_setup_state())

@app.route('/api/nexus/setup/step', methods=['POST'])
def api_nexus_setup_step():
    return jsonify(complete_setup_step(request.json or {}))

@app.route('/api/nexus/setup/finalize', methods=['POST'])
def api_nexus_setup_finalize():
    return jsonify(finalize_setup())

# ─── PC AGENT — EXECUTORES ────────────────────────────────────────────────────






@app.route('/api/tasks', methods=['GET'])
def api_tasks():
    return jsonify(get_all_tasks())

@app.route('/api/tasks/cancel', methods=['POST'])
def api_cancel_task():
    data    = request.json or {}
    task_id = data.get('id', '')
    success = cancel_task(task_id)
    return jsonify({'success': success})

@app.route('/api/news', methods=['GET'])
def api_news():
    """Busca notícias diretamente via GET."""
    topic = request.args.get('topic', '')
    count = int(request.args.get('count', 3))
    result = execute_get_news({'query': topic, 'count': count}, sid='api')
    return jsonify({'result': result, 'topics': NEWS_TOPICS})

@app.route('/api/trello/boards', methods=['GET'])
def api_trello_boards():
    """Lista boards do Trello."""
    import requests as req
    if not TRELLO_API_KEY or not TRELLO_TOKEN:
        return jsonify({'error': 'Trello não configurado'}), 503
    resp = req.get(
        'https://api.trello.com/1/members/me/boards',
        params={'key': TRELLO_API_KEY, 'token': TRELLO_TOKEN},
        timeout=10
    )
    return jsonify(resp.json())

@app.route('/api/asana/projects', methods=['GET'])
def api_asana_projects():
    """Lista projetos do Asana."""
    import requests as req
    if not ASANA_TOKEN:
        return jsonify({'error': 'Asana não configurado'}), 503
    resp = req.get(
        'https://app.asana.com/api/1.0/projects',
        headers={'Authorization': f'Bearer {ASANA_TOKEN}'},
        timeout=10
    )
    return jsonify(resp.json())

@app.route('/api/personalities', methods=['GET'])
def api_personalities():
    return jsonify(get_all_personalities())

@app.route('/api/personality/set', methods=['POST'])
def api_set_personality():
    data = request.json or {}
    name = data.get('personality', 'jarvis')
    if set_personality(name):
        p = get_personality()
        socketio.emit('personality_changed', {
            'id':    name,
            'name':  p['name'],
            'emoji': p['emoji'],
            'color': p['color'],
            'accent': p['accent'],
        })
        return jsonify({'success': True, 'personality': p})
    return jsonify({'success': False, 'error': 'Personalidade não encontrada'}), 404

# ─── ROTAS TEMPORAL LOBE ──────────────────────────────────────────────────────

@app.route('/api/temporal/timeline', methods=['GET'])
def api_temporal_timeline():
    """Retorna memórias e eventos de um período."""
    if not TEMPORAL_AVAILABLE:
        return jsonify({'error': 'Temporal Lobe indisponível'}), 503
    period = request.args.get('period', 'hoje')
    try:
        data = _tl.get_memories_by_period(period)
        return jsonify(data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/temporal/reminders', methods=['GET'])
def api_temporal_reminders():
    """Lista lembretes pendentes."""
    if not TEMPORAL_AVAILABLE:
        return jsonify([])
    try:
        reminders = _tl.get_all_reminders()
        return jsonify(reminders)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/temporal/reminders/add', methods=['POST'])
def api_temporal_reminders_add():
    """Adiciona um lembrete manualmente."""
    if not TEMPORAL_AVAILABLE:
        return jsonify({'success': False}), 503
    data = request.json or {}
    content    = data.get('content', '')
    remind_at  = data.get('remind_at', '')
    if not content or not remind_at:
        return jsonify({'success': False, 'error': 'content e remind_at obrigatórios'}), 400
    try:
        reminder = _tl.add_reminder(content, remind_at)
        return jsonify({'success': True, 'reminder': reminder})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/temporal/reminders/delete', methods=['POST'])
def api_temporal_reminders_delete():
    """Remove um lembrete pelo ID."""
    if not TEMPORAL_AVAILABLE:
        return jsonify({'success': False}), 503
    data = request.json or {}
    rid = data.get('id')
    try:
        _tl.delete_reminder(rid)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/temporal/episodes', methods=['GET'])
def api_temporal_episodes():
    """Lista episódios de memória episódica."""
    if not TEMPORAL_AVAILABLE:
        return jsonify([])
    limit = int(request.args.get('limit', 30))
    mood  = request.args.get('mood', None)
    try:
        episodes = _tl.get_recent_episodes(limit=limit, mood=mood if mood and mood != 'todos' else None)
        return jsonify(episodes)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/temporal/context', methods=['GET'])
def api_temporal_context():
    """Retorna contexto geo-temporal atual."""
    if not TEMPORAL_AVAILABLE:
        return jsonify({'error': 'Temporal Lobe indisponível'}), 503
    tz_offset = int(request.args.get('tz_offset', -3))
    try:
        ctx = _tl.get_geo_temporal_context(timezone_offset=tz_offset)
        return jsonify(ctx)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/temporal/decay/run', methods=['POST'])
def api_temporal_decay_run():
    """Executa o Temporal Decay manualmente."""
    if not TEMPORAL_AVAILABLE:
        return jsonify({'success': False}), 503
    try:
        removed = _tl.run_decay_cleanup()
        return jsonify({'success': True, 'removed': removed, 'message': 'Temporal Decay executado'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ─── PC AGENT — ROTAS REST ────────────────────────────────────────────────────

@app.route('/api/pc_agent/status', methods=['GET'])
def api_pc_agent_status():
    """Status e disponibilidade do PC Agent."""
    return jsonify(get_agent_status())

@app.route('/api/pc_agent/screenshot', methods=['GET'])
def api_pc_agent_screenshot():
    """Retorna screenshot atual da tela como base64."""
    if not PC_AGENT_AVAILABLE:
        return jsonify({'error': 'PC Agent indisponível'}), 503
    try:
        b64, meta = capture_screen_b64()
        return jsonify({'screenshot_b64': b64, 'meta': meta})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/pc_agent/analyze', methods=['POST'])
def api_pc_agent_analyze():
    """Análise rápida da tela via POST."""
    if not PC_AGENT_AVAILABLE:
        return jsonify({'error': 'PC Agent indisponível'}), 503
    data  = request.json or {}
    query = data.get('query', 'Descreva o que está na tela.')
    try:
        result = quick_screen_analysis(query, gemini_client)
        audio  = generate_tts(result)
        return jsonify({'result': result, 'audio_b64': audio})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/pc_agent/run', methods=['POST'])
def api_pc_agent_run():
    """Executa uma tarefa autônoma via REST (sem SocketIO)."""
    if not PC_AGENT_AVAILABLE:
        return jsonify({'error': 'PC Agent indisponível'}), 503
    data = request.json or {}
    task = data.get('task', '')
    if not task:
        return jsonify({'error': 'Campo "task" obrigatório'}), 400
    result = run_pc_agent(task=task, gemini_client=gemini_client, max_iterations=5)
    return jsonify(result)

# ─── EVENTOS SOCKETIO ─────────────────────────────────────────────────────────
@socketio.on('connect')
def on_connect():
    sid = request.sid
    chat_sessions[sid] = []
    print(f'[JARVIS] Cliente conectado: {sid}')
    emit('status_update', {'step': 'connected', 'message': 'Conectado'})

    # Envia perfil do usuário para o frontend personalizar a interface
    from memory import get_profile
    profile = get_profile()
    emit('profile_loaded', profile)

@socketio.on('disconnect')
def on_disconnect():
    sid = request.sid
    chat_sessions.pop(sid, None)
    print(f'[JARVIS] Cliente desconectado: {sid}')

@socketio.on('confirm_response')
def on_confirm_response(data):
    """Recebe a resposta do overlay de confirmação do frontend."""
    confirm_id = data.get('id', '')
    confirmed  = data.get('confirmed', False)
    if confirm_id in _pending_confirms:
        _confirm_results[confirm_id] = confirmed
        _pending_confirms[confirm_id].set()
        print(f'[JARVIS] Confirmação {confirm_id[:8]}: {"SIM" if confirmed else "NÃO"}')

@socketio.on('greeting_request')
def on_greeting():
    sid = request.sid
    hora = datetime.datetime.now().hour

    if 5 <= hora < 12:
        periodo, saudacao = 'manhã', 'Bom dia'
    elif 12 <= hora < 18:
        periodo, saudacao = 'tarde', 'Boa tarde'
    else:
        periodo, saudacao = 'noite', 'Boa noite'

    # Personaliza a saudação com o nome do usuário se disponível
    user_name = get_profile_field('user_name')
    if user_name:
        texto = f'{saudacao}, {user_name}, Senhor. JARVIS está operacional e a seu serviço. Todos os sistemas funcionando normalmente nesta {periodo}.'
    else:
        texto = f'{saudacao}, Senhor. JARVIS está operacional e a seu serviço. Todos os sistemas funcionando normalmente nesta {periodo}.'

    def _run():
        try:
            socketio.emit('status_update', {'step': 'speaking', 'message': 'Iniciando...'}, room=sid)
            audio_b64 = generate_tts(texto)
            socketio.emit('jarvis_response', {
                'text': texto,
                'audio_b64': audio_b64,
                'api_used': 'system',
            }, room=sid)
        except Exception as e:
            print(f'[JARVIS] Erro na saudação: {e}')

    Thread(target=_run, daemon=True).start()

@socketio.on('wake_word_greeting')
def on_wake_word_greeting():
    """
    Saudação especial ativada pela frase de wake word.
    Chamada pelo frontend quando detecta ?wake=1 na URL.
    """
    sid = request.sid
    hora = datetime.datetime.now().hour

    # Saudação baseada no horário
    if 5 <= hora < 12:
        periodo = 'manhã'
        saudacao_base = 'Bom dia'
    elif 12 <= hora < 18:
        periodo = 'tarde'
        saudacao_base = 'Boa tarde'
    else:
        periodo = 'noite'
        saudacao_base = 'Boa noite'

    # Frases personalizadas por período do dia
    if 5 <= hora < 12:
        texto = (
            'Bom dia Senhor. Ouvi seu chamado e estou completamente operacional. '
            'Está um ótimo dia. Todos os sistemas foram inicializados com sucesso nesta manhã.'
        )
    elif 12 <= hora < 18:
        texto = (
            'Boa tarde Senhor. Ouvi seu chamado e estou completamente operacional. '
            'Está uma ótima tarde. Todos os sistemas foram inicializados com sucesso nesta tarde.'
        )
    else:
        texto = (
            'Boa noite Senhor. Ouvi seu chamado e estou completamente operacional. '
            'Está uma ótima noite. Todos os sistemas foram inicializados com sucesso nesta noite.'
        )

    def _run():
        try:
            socketio.emit('status_update', {
                'step': 'speaking',
                'message': 'Wake word ativada!'
            }, room=sid)
            audio_b64 = generate_tts(texto)
            socketio.emit('jarvis_response', {
                'text': texto,
                'audio_b64': audio_b64,
                'api_used': 'wake_word'
            }, room=sid)
        except Exception as e:
            print(f'[JARVIS] Erro na saudação wake word: {e}')

    Thread(target=_run, daemon=True).start()

@socketio.on('clap_detected')
def on_clap_detected(data=None):
    """
    Disparado pelo frontend quando detecta palma dupla.
    Executa o protocolo 'Hora do Papai Trabalhar' direto, sem passar pela IA.
    """
    sid = request.sid
    print(f'[JARVIS] 👏👏 Palma dupla detectada — disparando protocolo')

    def _run():
        try:
            # 1. Executar ação (Spotify)
            socketio.emit('status_update', {'step': 'executing', 'message': 'Protocolo ativado...'}, room=sid)
            action_context = execute_protocol_work_time({}, sid)

            # 2. Frase fixa
            resposta = 'Iniciando protocolo hora do papai trabalhar, bom trabalho Senhor.'

            # 3. Registrar no histórico e log permanente
            history = chat_sessions.get(sid, [])
            history.append({'role': 'user', 'content': '[palma dupla detectada]'})
            history.append({'role': 'assistant', 'content': resposta})
            chat_sessions[sid] = history[-20:]
            log_message('user', '[palma dupla detectada]', intent='protocol_work_time')
            log_message('assistant', resposta)

            # 4. Gerar áudio
            socketio.emit('status_update', {'step': 'speaking', 'message': 'Sintetizando voz...'}, room=sid)
            audio_b64 = generate_tts(resposta)

            # 5. Enviar resposta ao frontend
            socketio.emit('jarvis_response', {
                'text':      resposta,
                'audio_b64': audio_b64,
                'api_used':  'protocol',
                'intent':    'protocol_work_time',
            }, room=sid)

            print('[JARVIS] Protocolo executado com sucesso')

        except Exception as e:
            print(f'[JARVIS] Erro no protocolo: {e}')
            import traceback; traceback.print_exc()
            socketio.emit('error', {'message': f'Erro no protocolo: {str(e)}'}, room=sid)

    Thread(target=_run, daemon=True).start()


@socketio.on('user_message')
def on_user_message(data):
    sid  = request.sid
    text = data.get('text', '').strip()
    if not text:
        return

    print(f'[JARVIS] Mensagem: {text}')

    # Verifica se há uma pergunta pendente (fluxo conversacional)
    if sid in _pending_asks:
        ask_data = _pending_asks.pop(sid)
        context = ask_data.get('context', '')
        meta = ask_data.get('meta', {})

        if context == 'whatsapp_message':
            # O texto é a resposta à pergunta "Qual mensagem?"  
            phone = meta.get('phone', '')
            contact = meta.get('contact', '')
            def _send_whatsapp_reply():
                try:
                    result = execute_send_whatsapp({'phone': phone, 'contact': contact, 'message': text}, sid)
                    audio_b64 = generate_tts(result)
                    socketio.emit('jarvis_response', {
                        'text': result, 'audio_b64': audio_b64,
                        'api_used': 'system', 'intent': 'send_whatsapp'
                    }, room=sid)
                except Exception as e:
                    socketio.emit('error', {'message': str(e)}, room=sid)
            Thread(target=_send_whatsapp_reply, daemon=True).start()
            return

    def _process():
        try:
            # 0a. Verifica troca de personalidade por voz
            new_personality = detect_personality_change(text)
            if new_personality:
                old_name = get_current_name()
                set_personality(new_personality)
                p = get_personality()
                socketio.emit('personality_changed', {
                    'id':    new_personality,
                    'name':  p['name'],
                    'emoji': p['emoji'],
                    'color': p['color'],
                    'accent': p['accent'],
                }, room=sid)
                resposta = f'{p["emoji"]} Personalidade alterada para {p["name"]}, Senhor. Como posso servi-lo?'
                audio_b64 = generate_tts(resposta)
                socketio.emit('jarvis_response', {
                    'text': resposta, 'audio_b64': audio_b64,
                    'api_used': 'system', 'intent': 'set_personality',
                }, room=sid)
                return

            # 0b. Detecção de emoção — emite evento para o Orb imediatamente
            emotion_data = {}
            if EMOTION_AVAILABLE:
                try:
                    emotion_data = emit_emotion_event(text, socketio, sid)
                except Exception as _eme:
                    print(f'[JARVIS] Emotion Engine: {_eme}')

            # 0c. Comando Pomodoro — tratado antes do pipeline de intenção
            if POMODORO_AVAILABLE and is_pomodoro_command(text):
                mgr = get_pomodoro_manager()
                if mgr:
                    resposta = handle_pomodoro_command(text, mgr, sid)
                    audio_b64 = generate_tts(resposta)
                    socketio.emit('jarvis_response', {
                        'text': resposta, 'audio_b64': audio_b64,
                        'api_used': 'pomodoro', 'intent': 'pomodoro',
                    }, room=sid)
                    log_message('user', text, intent='pomodoro')
                    log_message('assistant', resposta)
                    return

            # 0d. Briefing diário
            if BRIEFING_AVAILABLE and is_briefing_request(text):
                socketio.emit('status_update', {'step': 'thinking', 'message': 'Preparando briefing...'}, room=sid)
                deps = {
                    'groq_client': groq_client, 'GROQ_AVAILABLE': GROQ_AVAILABLE,
                    'gemini_client': gemini_client, 'GEMINI_AVAILABLE': GEMINI_AVAILABLE,
                    'OLLAMA_AVAILABLE': OLLAMA_AVAILABLE, 'OLLAMA_BASE_URL': OLLAMA_BASE_URL,
                    'OLLAMA_MODEL': OLLAMA_MODEL, 'GEMINI_MODELS': GEMINI_MODELS,
                    'get_all_tasks': get_all_tasks,
                    'TRELLO_API_KEY': TRELLO_API_KEY, 'ASANA_TOKEN': ASANA_TOKEN,
                }
                resposta  = generate_briefing(deps)
                audio_b64 = generate_tts(resposta)
                socketio.emit('jarvis_response', {
                    'text': resposta, 'audio_b64': audio_b64,
                    'api_used': 'briefing', 'intent': 'briefing',
                }, room=sid)
                log_message('user', text, intent='briefing')
                log_message('assistant', resposta)
                return

            # 1. Detectar intenção
            socketio.emit('status_update', {'step': 'thinking', 'message': 'Analisando...'}, room=sid)
            intent_data = detect_intent(text)

            # 2. Executar ação se necessário
            action_context = None
            if intent_data.get('intent') not in ('conversation', None):
                action_context = dispatch_intent(intent_data, sid)

                # Se a ação pediu uma pergunta ao usuário (fluxo conversacional)
                if action_context and action_context.startswith('__ASK__:'):
                    question = action_context[8:]
                    # Registra pergunta pendente
                    _pending_asks[sid] = {
                        'context': intent_data.get('intent', '').replace('send_', '') + '_message',
                        'meta': intent_data.get('params', {}),
                    }
                    # Envia pergunta ao frontend
                    audio_b64 = generate_tts(question)
                    socketio.emit('jarvis_response', {
                        'text': question, 'audio_b64': audio_b64,
                        'api_used': 'system', 'intent': 'ask',
                    }, room=sid)
                    return

            # 3. Extrair fatos da mensagem e salvar na memória (em background)
            Thread(target=extract_facts_via_ai, args=(text,), daemon=True).start()

            # 4. Montar histórico da sessão atual
            history = chat_sessions.get(sid, [])
            history.append({'role': 'user', 'content': text})
            if len(history) > 20:
                history = history[-20:]

            # ─── PROTOCOLO ESPECIAL: frase fixa, sem IA ───
            if intent_data.get('intent') == 'protocol_work_time':
                resposta = 'Iniciando protocolo hora do papai trabalhar.'
                api_usada = 'protocol'
            else:
                # 5. Gerar resposta IA (com SOUL + memórias + emoção + humor injetados)
                socketio.emit('status_update', {'step': 'thinking', 'message': 'Gerando resposta...'}, room=sid)
                emotion_name = emotion_data.get('emotion', '')
                resposta, api_usada = generate_ai_response(
                    history,
                    context=action_context or '',
                    emotion=emotion_name,
                )

            # 6. Salvar no histórico da sessão e no log permanente
            history.append({'role': 'assistant', 'content': resposta})
            chat_sessions[sid] = history

            log_message('user',      text,    intent=intent_data.get('intent'))
            log_message('assistant', resposta)

            # Salva no Mem0 em background (memória de longo prazo)
            if MEM0_AVAILABLE:
                Thread(target=mem0_add, args=(text, 'user'), daemon=True).start()
                Thread(target=mem0_add, args=(resposta, 'assistant'), daemon=True).start()

            # 7. Gerar áudio TTS
            socketio.emit('status_update', {'step': 'speaking', 'message': 'Sintetizando voz...'}, room=sid)
            audio_b64 = generate_tts(resposta)

            # 8. Enviar resposta ao frontend
            socketio.emit('jarvis_response', {
                'text':      resposta,
                'audio_b64': audio_b64,
                'api_used':  api_usada,
                'intent':    intent_data.get('intent', 'conversation'),
            }, room=sid)

            print(f'[JARVIS] Resposta via {api_usada}')

        except Exception as e:
            print(f'[JARVIS] Erro no processamento: {e}')
            import traceback; traceback.print_exc()
            socketio.emit('error', {'message': f'Erro interno: {str(e)}'}, room=sid)

    Thread(target=_process, daemon=True).start()

# ─── PC AGENT — EVENTOS SOCKETIO ──────────────────────────────────────────────

@socketio.on('pc_agent_run')
def on_pc_agent_run(data):
    """
    Evento SocketIO para execução do PC Agent com feedback em tempo real.
    Frontend envia: {task: 'tarefa em linguagem natural'}
    """
    sid  = request.sid
    task = (data or {}).get('task', '').strip()
    if not task:
        emit('error', {'message': 'Tarefa não especificada'})
        return

    print(f'[JARVIS] PC Agent task via SocketIO: {task}')

    def _run():
        try:
            result = execute_pc_agent_task({'task': task}, sid)
            audio_b64 = generate_tts(result)
            socketio.emit('jarvis_response', {
                'text':      result,
                'audio_b64': audio_b64,
                'api_used':  'pc_agent',
                'intent':    'pc_agent_task',
            }, room=sid)
        except Exception as e:
            socketio.emit('error', {'message': f'Erro no PC Agent: {str(e)}'}, room=sid)

    Thread(target=_run, daemon=True).start()


@socketio.on('pc_agent_stop')
def on_pc_agent_stop():
    """Para o agente movendo o mouse para o canto superior esquerdo (failsafe pyautogui)."""
    try:
        import pyautogui
        pyautogui.moveTo(0, 0)
        emit('status_update', {'step': 'stopped', 'message': 'PC Agent interrompido'})
        print('[JARVIS] PC Agent interrompido via failsafe')
    except Exception as e:
        emit('error', {'message': f'Erro ao parar agente: {e}'})


# ─── EVENTOS SOCKETIO v2.3 ────────────────────────────────────────────────────

@socketio.on('confirm_response')
def on_confirm_response(data):
    """
    Resposta do usuário ao modal de confirmação (botão Sim/Não na UI).
    data: {id: str, confirmed: bool}
    """
    confirm_id = (data or {}).get('id', '')
    confirmed  = bool((data or {}).get('confirmed', False))
    if CONFIRM_V23_AVAILABLE:
        resolve_confirmation(confirm_id, confirmed)
    else:
        # Legado — compatibilidade com sistema antigo
        from App import _confirm_results, _pending_confirms
        if confirm_id in _pending_confirms:
            _confirm_results[confirm_id] = confirmed
            _pending_confirms[confirm_id].set()


@socketio.on('pomodoro_command')
def on_pomodoro_command(data):
    """
    Comando Pomodoro enviado diretamente pelo frontend (botões da UI).
    data: {command: 'start'|'pause'|'resume'|'skip'|'stop'|'status'}
    """
    sid = request.sid
    if not POMODORO_AVAILABLE:
        emit('error', {'message': 'Pomodoro indisponível'})
        return
    mgr = get_pomodoro_manager()
    if not mgr:
        emit('error', {'message': 'Pomodoro não inicializado'})
        return
    cmd = (data or {}).get('command', 'status')
    cmd_map = {
        'start':  mgr.start,
        'pause':  mgr.pause,
        'resume': mgr.resume,
        'skip':   mgr.skip,
        'stop':   mgr.stop,
        'status': mgr.get_status,
    }
    fn = cmd_map.get(cmd, mgr.get_status)
    try:
        if cmd == 'start':
            resposta = fn(sid)
        else:
            resposta = fn()
        audio_b64 = generate_tts(resposta)
        socketio.emit('jarvis_response', {
            'text': resposta, 'audio_b64': audio_b64,
            'api_used': 'pomodoro', 'intent': 'pomodoro',
        }, room=sid)
    except Exception as e:
        emit('error', {'message': f'Erro no Pomodoro: {e}'})


@socketio.on('briefing_request')
def on_briefing_request(data=None):
    """Solicita briefing diário diretamente pelo frontend."""
    sid = request.sid
    def _run():
        if not BRIEFING_AVAILABLE:
            socketio.emit('jarvis_response', {
                'text': 'Briefing indisponível no momento, Senhor.',
                'audio_b64': generate_tts('Briefing indisponível no momento, Senhor.'),
                'api_used': 'system', 'intent': 'briefing',
            }, room=sid)
            return
        deps = {
            'groq_client': groq_client, 'GROQ_AVAILABLE': GROQ_AVAILABLE,
            'gemini_client': gemini_client, 'GEMINI_AVAILABLE': GEMINI_AVAILABLE,
            'OLLAMA_AVAILABLE': OLLAMA_AVAILABLE, 'OLLAMA_BASE_URL': OLLAMA_BASE_URL,
            'OLLAMA_MODEL': OLLAMA_MODEL, 'GEMINI_MODELS': GEMINI_MODELS,
            'get_all_tasks': get_all_tasks,
            'TRELLO_API_KEY': TRELLO_API_KEY, 'ASANA_TOKEN': ASANA_TOKEN,
        }
        socketio.emit('status_update', {'step': 'thinking', 'message': 'Preparando briefing...'}, room=sid)
        resposta  = generate_briefing(deps)
        audio_b64 = generate_tts(resposta)
        socketio.emit('jarvis_response', {
            'text': resposta, 'audio_b64': audio_b64,
            'api_used': 'briefing', 'intent': 'briefing',
        }, room=sid)
    Thread(target=_run, daemon=True).start()


# ─── ROTAS API v2.3 ───────────────────────────────────────────────────────────

@app.route('/api/v23/router', methods=['GET'])
def api_router_status():
    """Status em tempo real do LLM Router: circuit breakers, cache stats."""
    try:
        return jsonify(get_router().status())
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/v23/router/reset', methods=['POST'])
def api_router_reset():
    """Reseta circuit breakers (admin/debug)."""
    try:
        get_router().reset_circuits()
        get_router().clear_intent_cache()
        return jsonify({'success': True, 'message': 'Router resetado, Senhor.'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/v23/status', methods=['GET'])
def api_v23_status():
    """Status de todos os módulos v2.3."""
    try:
        from pomodoro import get_pomodoro_manager
        pomodoro_state = get_pomodoro_manager().get_status_dict() if (POMODORO_AVAILABLE and get_pomodoro_manager()) else None
    except Exception:
        pomodoro_state = None

    return jsonify({
        'version':          '2.3',
        'codename':         'Consciência Avançada',
        'llm_router':       get_router().status(),
        'emotion_engine':   EMOTION_AVAILABLE,
        'briefing':         BRIEFING_AVAILABLE,
        'pomodoro':         POMODORO_AVAILABLE,
        'pomodoro_state':   pomodoro_state,
        'humor_contextual': HUMOR_AVAILABLE,
        'wake_word':        get_wake_word_status(),
        'confirmation_v23': CONFIRM_V23_AVAILABLE,
        'self_reflection':  REFLECTION_AVAILABLE,
    })

@app.route('/api/v23/briefing', methods=['GET'])
def api_briefing():
    """Gera briefing via HTTP (para uso externo ou debug)."""
    deps = {
        'groq_client': groq_client, 'GROQ_AVAILABLE': GROQ_AVAILABLE,
        'gemini_client': gemini_client, 'GEMINI_AVAILABLE': GEMINI_AVAILABLE,
        'OLLAMA_AVAILABLE': OLLAMA_AVAILABLE, 'OLLAMA_BASE_URL': OLLAMA_BASE_URL,
        'OLLAMA_MODEL': OLLAMA_MODEL, 'GEMINI_MODELS': GEMINI_MODELS,
        'get_all_tasks': get_all_tasks,
        'TRELLO_API_KEY': TRELLO_API_KEY, 'ASANA_TOKEN': ASANA_TOKEN,
    }
    texto = generate_briefing(deps)
    return jsonify({'text': texto, 'success': True})

@app.route('/api/v23/pomodoro', methods=['GET'])
def api_pomodoro_status():
    """Status atual do Pomodoro."""
    if not POMODORO_AVAILABLE or not get_pomodoro_manager():
        return jsonify({'available': False})
    return jsonify(get_pomodoro_manager().get_status_dict())

@app.route('/api/v23/emotion', methods=['POST'])
def api_emotion_detect():
    """Detecta emoção de um texto (para debug/frontend)."""
    if not EMOTION_AVAILABLE:
        return jsonify({'available': False}), 503
    from emotion_engine import detect_emotion
    data = request.get_json(force=True) or {}
    text = data.get('text', '')
    return jsonify(detect_emotion(text))

@app.route('/api/v23/humor_context', methods=['GET'])
def api_humor_context():
    """Retorna contexto de humor atual."""
    if not HUMOR_AVAILABLE:
        return jsonify({'available': False})
    from humor_contextual import get_context_info
    return jsonify(get_context_info())


# ─── INICIALIZAÇÃO ────────────────────────────────────────────────────────────
_init_skills()

if __name__ == '__main__':
    print("""
╔══════════════════════════════════════════════════════════════╗
║        J.A.R.V.I.S. — SISTEMA INICIALIZADO                  ║
║   Just A Rather Very Intelligent System v2.3                 ║
║   "Consciência Avançada"                                     ║
║   Wake Word · Emoção · Briefing · Pomodoro · Humor           ║
╠══════════════════════════════════════════════════════════════╣
║  Acesse: http://localhost:5000                               ║
╚══════════════════════════════════════════════════════════════╝
    """)
    # Inicia módulos de ciclo de vida
    init_nexus()
    init_scheduler(socketio, generate_tts)

    # Inicia Pomodoro
    if POMODORO_AVAILABLE:
        init_pomodoro(socketio, tts_fn=generate_tts)
        print('[JARVIS] Pomodoro Engine: iniciado')

    # Inicia Wake Word (em thread separada, não bloqueia o Flask)
    if WAKE_WORD_AVAILABLE:
        _ww_engine = init_wake_word(socketio, tts_fn=generate_tts)
        if _ww_engine and _ww_engine.is_available:
            print('[JARVIS] Wake Word Engine: escutando microfone')
        else:
            print('[JARVIS] Wake Word Engine: módulos opcionais ausentes (pip install faster-whisper sounddevice)')

    # Inicia Self Reflection
    if REFLECTION_AVAILABLE:
        try:
            from neural_core import make_simple_ai_caller
            _ai_simple = make_simple_ai_caller(gemini_client, groq_client, GEMINI_MODELS)
            _reflector  = init_self_reflection(_ai_simple, generate_tts, socketio)
            print('[JARVIS] Self Reflection: pronto')
        except Exception as _re:
            print(f'[JARVIS] Self Reflection: erro na inicialização ({_re})')

    socketio.run(app, host='0.0.0.0', port=5000, debug=False, allow_unsafe_werkzeug=True)